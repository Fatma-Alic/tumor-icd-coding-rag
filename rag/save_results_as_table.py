import glob
import os
import re
from typing import Dict, List, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


# =============================================================================
# CONFIG
# =============================================================================
ROOT_RESULTS_DIR_ICD10 = r"/home/alic/LLIT-RAG/ICD10"
ROOT_RESULTS_DIR_ICDO = r"/home/alic/LLIT-RAG/ICDO"

NAMES_FILE = r"/home/alic/LLIT-RAG/names_embedding_models_for_paper.csv"

BASELINE_SIM_ROOT_ICD10 = r"/home/alic/ChromaDB/results_for_table/ChromaDB_icd10/"
BASELINE_SIM_ROOT_ICDO = r"/home/alic/ChromaDB/results_for_table/ChromaDB_icdo/"
BASELINE_CLF_ROOT_ICD10 = r"/home/alic/ChromaDB/results_for_table/Classifier_icd10/"
BASELINE_CLF_ROOT_ICDO = r"/home/alic/ChromaDB/results_for_table/Classifier_icdo/"

OUTPUT_XLSX = r"/home/alic/LLIT-RAG/overall_results.xlsx"

ALLOWED_METHOD_ROOTS_ICD10 = {
    "ChromaDB/Tumordiagnose/Prompt_check",
    "Classifier/Tumordiagnose/Prompt_check",
    "Zero-shot-prompting/Tumordiagnose/Prompt_check",
    "Random/Tumordiagnose/Prompt_check",
}

ALLOWED_METHOD_ROOTS_ICDO = {
    "ChromaDB/Label",
    "Classifier/Label",
    "Zero-shot-prompting/Label",
    "Random/Label",
}

METHOD_FOLDER_MAP = {
    "ChromaDB/Tumordiagnose/Prompt_check": "RAG-prompting similarity search",
    "Classifier/Tumordiagnose/Prompt_check": "RAG-prompting classifier head",
    "Zero-shot-prompting/Tumordiagnose/Prompt_check": "zero-shot-prompting",
    "Random/Tumordiagnose/Prompt_check": "Three-shot-prompting",
    "ChromaDB/Label": "RAG-prompting similarity search",
    "Classifier/Label": "RAG-prompting classifier head",
    "Zero-shot-prompting/Label": "zero-shot-prompting",
    "Random/Label": "Three-shot-prompting",
}

LLM_MODEL_MAP = {
    "Llama-3.1-8B": "Llama 3.1",
    "Lama-3.1-8B": "Llama 3.1",
    "Llama-3.3-70B": "Llama 3.3",
    "Llama 3.3-70B": "Llama 3.3",
}

LLM_VARIANT_MAP = {
    "LLM_base": "base",
    "LLM_peft": "PEFT",
    "Llama_base": "base",
    "Llama_peft": "PEFT",
    "base": "base",
    "peft": "PEFT",
}

LLM_ORDER = [
    "Llama 3.1 base",
    "Llama 3.1 PEFT",
    "Llama 3.3 base",
    "Llama 3.3 PEFT",
]

METHOD_ORDER = [
    "retrieval-baseline: similarity search",
    "retrieval-baseline: classifier head",
    "zero-shot-prompting",
    "Three-shot-prompting",
    "RAG-prompting similarity search",
    "RAG-prompting classifier head",
]

METRIC_COLUMNS = [
    "Malformed",
    "Accuracy",
    "Partial Accuracy",
    "F1 weighted (exact)",
    "F1 weighted (partial)",
    "F1 macro (exact)",
    "F1 macro (partial)",
]

COLS_ANALYSIS = {
    "Accuracy": "Accuracy",
    "Partial Accuracy": "Partial Accuracy",
    "F1 weighted (exact)": "F1 Exact (weighted)",
    "F1 weighted (partial)": "F1 Partial (weighted)",
    "F1 macro (exact)": "F1 Exact (macro)",
    "F1 macro (partial)": "F1 Partial (macro)",
}
ANALYSIS_NUM_Q = "Num Questions"
ANALYSIS_INVALID = "Invalid Answer"

COLS_SUMMARY = {
    "Accuracy": "accuracy_exact",
    "Partial Accuracy": "accuracy_partial",
    "F1 weighted (exact)": "f1_weighted_exact",
    "F1 weighted (partial)": "f1_weighted_partial",
    "F1 macro (exact)": "f1_macro_exact",
    "F1 macro (partial)": "f1_macro_partial",
}
SUMMARY_N = "n_evaluated"
SUMMARY_INVALID = "invalid_pred_counted_wrong"


# =============================================================================
# HELPERS
# =============================================================================
def norm_key(s: str) -> str:
    """
    Normalize a string so it can be used as a matching key.

    Parameters:
        s (str): Input text.

    Returns:
        str: Normalized key.
    """
    s = "" if s is None else str(s)
    s = s.strip()
    s = s.replace("\\", "/")
    s = s.split("/")[-1]
    s = re.sub(r"\.(csv|CSV|xlsx|xls)$", "", s)

    s = re.sub(r"^(analysis_)", "", s, flags=re.I)
    s = re.sub(r"^(results?_)", "", s, flags=re.I)
    s = re.sub(r"^(result-)", "", s, flags=re.I)

    s = re.sub(r"_with_F1$", "", s, flags=re.I)
    s = re.sub(r"(_classifier.*)$", "", s, flags=re.I)
    s = re.sub(r"(_chromadb.*)$", "", s, flags=re.I)
    s = re.sub(r"(_predictions.*)$", "", s, flags=re.I)
    s = re.sub(r"(_shots_questions_\d+)$", "", s, flags=re.I)
    s = re.sub(r"(_questions_\d+)$", "", s, flags=re.I)
    s = re.sub(r"(_shots_\d+)$", "", s, flags=re.I)
    s = re.sub(r"(_icd10|_icdo|icd10|icdo)$", "", s, flags=re.I)

    s = s.strip("_- ")
    s = s.replace("–", "-").replace("—", "-")
    s = s.replace("-", "_")
    s = re.sub(r"_+", "_", s)
    s = re.sub(r"\s+", " ", s)
    return s.lower()


def normalize_embedding_id(s: str) -> str:
    """
    Clean an embedding file or folder name.

    Parameters:
        s (str): Raw embedding name.

    Returns:
        str: Cleaned embedding name.
    """
    s = "" if s is None else str(s)
    s = s.strip().replace("\\", "/")
    s = s.split("/")[-1]

    s = re.sub(r"\.(csv|CSV)$", "", s)
    s = re.sub(r"^(analysis_)", "", s, flags=re.I)
    s = re.sub(r"^(results?_)", "", s, flags=re.I)
    s = re.sub(r"^(result-)", "", s, flags=re.I)
    s = re.sub(r"_with_F1$", "", s, flags=re.I)
    s = re.sub(r"(_classifier.*)$", "", s, flags=re.I)
    s = re.sub(r"(_chromadb.*)$", "", s, flags=re.I)
    s = re.sub(r"(_predictions.*)$", "", s, flags=re.I)
    s = re.sub(r"(_icd10|_icdo)$", "", s, flags=re.I)

    s = s.strip("_- ")
    return s


def load_embedding_name_map(path: str) -> dict:
    """
    Load the mapping from raw model names to paper names.

    Parameters:
        path (str): Path to the CSV or Excel file.

    Returns:
        dict: Mapping from normalized name to paper name.
    """
    df = pd.read_excel(path) if path.lower().endswith(".xlsx") else pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]

    if "Paper_Name" not in df.columns:
        raise ValueError(
            f"'Paper_Name' not found in {path}. Columns: {df.columns.tolist()}"
        )

    map_out = {}

    for _, row in df.iterrows():
        paper_name = str(row.get("Paper_Name", "")).strip()
        if not paper_name or paper_name.lower() == "nan":
            continue

        map_out[norm_key(paper_name)] = paper_name

        for col in df.columns:
            if col == "Paper_Name":
                continue

            val = str(row.get(col, "")).strip()
            if not val or val.lower() == "nan":
                continue

            map_out[norm_key(val)] = paper_name

            cleaned = normalize_embedding_id(val)
            if cleaned:
                map_out[norm_key(cleaned)] = paper_name

    return map_out


def resolve_embedding_model_name(raw_embedding_id: str, name_map: dict) -> str:
    """
    Resolve a raw embedding name to the paper name.

    Parameters:
        raw_embedding_id (str): Raw embedding identifier.
        name_map (dict): Mapping from raw names to paper names.

    Returns:
        str: Resolved embedding name.
    """
    raw_embedding_id = "" if raw_embedding_id is None else str(raw_embedding_id).strip()
    cleaned = normalize_embedding_id(raw_embedding_id)

    candidates = [
        raw_embedding_id,
        cleaned,
        re.sub(r"(_shots_questions_\d+)$", "", cleaned, flags=re.I),
        re.sub(r"(_questions_\d+)$", "", cleaned, flags=re.I),
        re.sub(r"(_shots_\d+)$", "", cleaned, flags=re.I),
    ]

    for cand in candidates:
        key = norm_key(cand)
        if key in name_map:
            return name_map[key]

    return cleaned if cleaned else "unknown-embedding"


def to_percent_number(val):
    """
    Convert a numeric value to a percent value.

    Parameters:
        val: Input value.

    Returns:
        float | None: Rounded percent value or None.
    """
    try:
        x = float(val)
    except Exception:
        return None

    x = x if x > 1.5 else x * 100.0
    return round(x, 2)


def safe_float(val):
    """
    Safely convert a value to float.

    Parameters:
        val: Input value.

    Returns:
        float | None: Float value or None.
    """
    try:
        return float(val)
    except Exception:
        return None


def to_percent_from_count(count_val, total_val):
    """
    Calculate percentage from count and total.

    Parameters:
        count_val: Count value.
        total_val: Total value.

    Returns:
        float | None: Percentage or None.
    """
    c = safe_float(count_val)
    t = safe_float(total_val)
    if c is None or t is None or t == 0:
        return None
    return round((c / t) * 100.0, 2)


def detect_method_from_path(path: str) -> str:
    """
    Detect the method label from a file path.

    Parameters:
        path (str): File path.

    Returns:
        str: Method label.
    """
    norm_path = path.replace("\\", "/")
    for key, value in METHOD_FOLDER_MAP.items():
        if key in norm_path:
            return value
    return "unknown-method"


def detect_llm_from_path(path: str) -> str:
    """
    Detect the LLM and variant from a file path.

    Parameters:
        path (str): File path.

    Returns:
        str: LLM label.
    """
    parts = path.replace("\\", "/").split("/")
    llm_model = None
    llm_variant = None

    for p in parts:
        if p in LLM_MODEL_MAP:
            llm_model = LLM_MODEL_MAP[p]
        if p in LLM_VARIANT_MAP:
            llm_variant = LLM_VARIANT_MAP[p]

    if not llm_model or not llm_variant:
        return "unknown-llm"

    return f"{llm_model} {llm_variant}"


def extract_embedding_id_from_analysis_filename(filename: str) -> str:
    """
    Extract the embedding id from an analysis filename.

    Parameters:
        filename (str): File name.

    Returns:
        str: Extracted embedding id.
    """
    base = os.path.basename(filename)

    match = re.match(r"analysis_(.*)_with_F1\.csv$", base)
    if match:
        return match.group(1)

    match_2 = re.match(r"analysis_(.*)\.csv$", base)
    if match_2:
        return match_2.group(1)

    return base


def read_last_epoch_metrics_from_analysis(csv_path: str) -> dict:
    """
    Read the metrics from the last row of an analysis CSV.

    Parameters:
        csv_path (str): Path to the analysis CSV.

    Returns:
        dict: Metrics dictionary.
    """
    df = pd.read_csv(csv_path)
    df.columns = [c.strip() for c in df.columns]

    if df.empty:
        return {k: None for k in METRIC_COLUMNS}

    row = df.iloc[-1].to_dict()

    out = {}
    out["Malformed"] = to_percent_from_count(
        row.get(ANALYSIS_INVALID),
        row.get(ANALYSIS_NUM_Q),
    )
    for out_col, src_col in COLS_ANALYSIS.items():
        out[out_col] = to_percent_number(row.get(src_col))

    return out


def read_csv_anysep(path: str) -> pd.DataFrame:
    """
    Read a CSV file with comma or semicolon separator.

    Parameters:
        path (str): Path to the CSV file.

    Returns:
        pd.DataFrame: Loaded dataframe.
    """
    df = pd.read_csv(path, sep=",")
    if len(df.columns) == 1:
        df = pd.read_csv(path, sep=";")
    df.columns = [c.strip() for c in df.columns]
    return df


def extract_embedding_id_from_summary_row(row: dict) -> str:
    """
    Extract embedding id from one summary row.

    Parameters:
        row (dict): One row as dictionary.

    Returns:
        str: Extracted embedding id.
    """
    cand = str(row.get("file", "")).strip()
    if not cand:
        cand = str(row.get("path", "")).strip()
    if not cand:
        cand = str(row.get("source", "")).strip()

    emb_id = normalize_embedding_id(cand)
    return emb_id or "unknown-embedding"


def iter_all_summary_files(root_dir: str, prefix: str) -> list[str]:
    """
    Find all summary CSV files with a given prefix.

    Parameters:
        root_dir (str): Root directory to search in.
        prefix (str): File prefix.

    Returns:
        list[str]: List of matching file paths.
    """
    out = []
    prefix_low = prefix.lower()

    for dirpath, _, filenames in os.walk(root_dir):
        for filename in filenames:
            filename_low = filename.lower()
            if filename_low.startswith(prefix_low) and filename_low.endswith(".csv"):
                out.append(os.path.join(dirpath, filename))

    return sorted(out)


def iter_all_analysis_files(
    root_dir: str,
    allowed_method_roots: Set[str],
) -> list[str]:
    """
    Find all analysis CSV files in allowed method folders.

    Parameters:
        root_dir (str): Root directory to search in.
        allowed_method_roots (Set[str]): Allowed method folder patterns.

    Returns:
        list[str]: List of matching file paths.
    """
    found = set()

    for file_path in glob.glob(
        os.path.join(root_dir, "**", "analysis_*_with_F1.csv"),
        recursive=True,
    ):
        norm = file_path.replace("\\", "/")
        if any(root in norm for root in allowed_method_roots):
            found.add(os.path.abspath(file_path))

    llama_folder_names = {
        "Llama-3.1-8B",
        "Lama-3.1-8B",
        "Llama-3.3-70B",
        "Llama 3.3-70B",
    }

    variant_names = (
        "LLM_base",
        "LLM_peft",
        "base",
        "peft",
        "Llama_base",
        "Llama_peft",
    )

    for dirpath, _, _ in os.walk(root_dir):
        current = os.path.basename(dirpath)

        if current in llama_folder_names:
            for variant in variant_names:
                results_dir = os.path.join(dirpath, variant, "Filtered", "results")
                if os.path.isdir(results_dir):
                    for file_path in glob.glob(
                        os.path.join(results_dir, "**", "analysis_*_with_F1.csv"),
                        recursive=True,
                    ):
                        norm = file_path.replace("\\", "/")
                        if any(root in norm for root in allowed_method_roots):
                            found.add(os.path.abspath(file_path))

    return sorted(found)


def deduplicate_results(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove duplicate rows from the results dataframe.

    Parameters:
        df (pd.DataFrame): Results dataframe.

    Returns:
        pd.DataFrame: Deduplicated dataframe.
    """
    dedup_cols = [
        "Code system",
        "Method",
        "LLM",
        "Embedding model",
        "Accuracy",
        "Partial Accuracy",
        "F1 weighted (exact)",
        "F1 weighted (partial)",
        "F1 macro (exact)",
        "F1 macro (partial)",
        "Malformed",
    ]

    for col in dedup_cols:
        if col not in df.columns:
            df[col] = None

    return df.drop_duplicates(subset=dedup_cols, keep="first").reset_index(drop=True)


def sort_results_for_paper(df: pd.DataFrame) -> pd.DataFrame:
    """
    Sort the dataframe in the order needed for the paper.

    Parameters:
        df (pd.DataFrame): Results dataframe.

    Returns:
        pd.DataFrame: Sorted dataframe.
    """
    df = df.copy()

    df["Method"] = pd.Categorical(
        df["Method"],
        categories=METHOD_ORDER,
        ordered=True,
    )
    df["LLM_sort"] = df["LLM"].map(
        {value: i for i, value in enumerate(LLM_ORDER)}
    ).fillna(999)

    df = df.sort_values(
        ["Code system", "Method", "Embedding model", "LLM_sort", "LLM"],
        na_position="last",
    ).reset_index(drop=True)

    df = df.drop(columns=["LLM_sort"])
    return df


def get_best_accuracy_row_indices(df: pd.DataFrame) -> set:
    """
    Find rows with the highest Accuracy per code system and LLM.

    Parameters:
        df (pd.DataFrame): Results dataframe.

    Returns:
        set: Dataframe indices of the best rows.
    """
    best_rows = set()

    df_tmp = df.copy()
    df_tmp = df_tmp[df_tmp["LLM"].isin(LLM_ORDER)].copy()
    df_tmp["Accuracy_num"] = pd.to_numeric(df_tmp["Accuracy"], errors="coerce")

    for (_, _), group in df_tmp.groupby(["Code system", "LLM"], dropna=False):
        group_valid = group.dropna(subset=["Accuracy_num"])
        if group_valid.empty:
            continue

        max_acc = group_valid["Accuracy_num"].max()
        winners = group_valid[group_valid["Accuracy_num"] == max_acc].index.tolist()
        best_rows.update(winners)

    return best_rows


# =============================================================================
# BUILD ONE CODE SYSTEM
# =============================================================================
def build_dataframe_single(
    code_system_label: str,
    root_results_dir: str,
    baseline_sim_root: str,
    baseline_clf_root: str,
    allowed_method_roots: Set[str],
) -> pd.DataFrame:
    """
    Build the results dataframe for one code system.

    Parameters:
        code_system_label (str): Label of the code system.
        root_results_dir (str): Root results directory.
        baseline_sim_root (str): Root folder for similarity-search baselines.
        baseline_clf_root (str): Root folder for classifier baselines.
        allowed_method_roots (Set[str]): Allowed method folder patterns.

    Returns:
        pd.DataFrame: Results dataframe for one code system.
    """
    name_map = load_embedding_name_map(NAMES_FILE)
    rows = []

    baseline_specs = [
        (
            "ChromaDB_metrics_summary",
            "retrieval-baseline: similarity search",
            baseline_sim_root,
        ),
        (
            "Classifier_metrics_summary",
            "retrieval-baseline: classifier head",
            baseline_clf_root,
        ),
    ]

    for prefix, method_label, base_root in baseline_specs:
        summary_files = iter_all_summary_files(base_root, prefix)

        if not summary_files:
            rows.append(
                {
                    "Code system": code_system_label,
                    "Method": method_label,
                    "Embedding model": "NOT FOUND",
                    "LLM": "—",
                    **{k: None for k in METRIC_COLUMNS},
                }
            )
            continue

        for summary_csv in summary_files:
            df_sum = read_csv_anysep(summary_csv)

            for _, row in df_sum.iterrows():
                row_dict = row.to_dict()

                raw_emb_id = extract_embedding_id_from_summary_row(row_dict)
                emb_model = resolve_embedding_model_name(raw_emb_id, name_map)

                metrics = {}
                metrics["Malformed"] = to_percent_from_count(
                    row_dict.get(SUMMARY_INVALID),
                    row_dict.get(SUMMARY_N),
                )
                for out_col, src_col in COLS_SUMMARY.items():
                    metrics[out_col] = to_percent_number(row_dict.get(src_col))

                rows.append(
                    {
                        "Code system": code_system_label,
                        "Method": method_label,
                        "LLM": "—",
                        "Embedding model": emb_model,
                        **metrics,
                    }
                )

    files = iter_all_analysis_files(root_results_dir, allowed_method_roots)

    for file_path in files:
        method = detect_method_from_path(file_path)
        llm = detect_llm_from_path(file_path)

        raw_emb_id = extract_embedding_id_from_analysis_filename(file_path)
        emb_model = resolve_embedding_model_name(raw_emb_id, name_map)

        metrics = read_last_epoch_metrics_from_analysis(file_path)

        rows.append(
            {
                "Code system": code_system_label,
                "Method": method,
                "LLM": llm,
                "Embedding model": emb_model,
                **metrics,
            }
        )

    df = pd.DataFrame(rows)

    visible_cols = [
        "Code system",
        "Method",
        "Embedding model",
        "LLM",
        "Accuracy",
        "Partial Accuracy",
        "F1 weighted (exact)",
        "F1 weighted (partial)",
        "F1 macro (exact)",
        "F1 macro (partial)",
        "Malformed",
    ]
    for col in visible_cols:
        if col not in df.columns:
            df[col] = None

    df = df[visible_cols]
    df = deduplicate_results(df)
    df = sort_results_for_paper(df)

    return df


# =============================================================================
# BUILD COMBINED
# =============================================================================
def build_dataframe_combined() -> pd.DataFrame:
    """
    Build the combined dataframe for ICD-10 and ICD-O-3.

    Parameters:
        None

    Returns:
        pd.DataFrame: Combined results dataframe.
    """
    df_icd10 = build_dataframe_single(
        code_system_label="ICD-10",
        root_results_dir=ROOT_RESULTS_DIR_ICD10,
        baseline_sim_root=BASELINE_SIM_ROOT_ICD10,
        baseline_clf_root=BASELINE_CLF_ROOT_ICD10,
        allowed_method_roots=ALLOWED_METHOD_ROOTS_ICD10,
    )

    df_icdo = build_dataframe_single(
        code_system_label="ICD-O-3",
        root_results_dir=ROOT_RESULTS_DIR_ICDO,
        baseline_sim_root=BASELINE_SIM_ROOT_ICDO,
        baseline_clf_root=BASELINE_CLF_ROOT_ICDO,
        allowed_method_roots=ALLOWED_METHOD_ROOTS_ICDO,
    )

    df_all = pd.concat([df_icd10, df_icdo], ignore_index=True)

    final_cols = [
        "Code system",
        "Method",
        "Embedding model",
        "LLM",
        "Accuracy",
        "Partial Accuracy",
        "F1 weighted (exact)",
        "F1 weighted (partial)",
        "F1 macro (exact)",
        "F1 macro (partial)",
        "Malformed",
    ]
    for col in final_cols:
        if col not in df_all.columns:
            df_all[col] = None

    df_all = df_all[final_cols]
    df_all = deduplicate_results(df_all)
    df_all = sort_results_for_paper(df_all)

    return df_all


# =============================================================================
# EXPORT
# =============================================================================
def write_excel(df: pd.DataFrame, output_path: str):
    """
    Write the dataframe to Excel and bold the best Accuracy values.

    Parameters:
        df (pd.DataFrame): Results dataframe.
        output_path (str): Output Excel file path.

    Returns:
        None
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    df.to_excel(output_path, index=False)

    best_idx = get_best_accuracy_row_indices(df)

    wb = load_workbook(output_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    accuracy_col_idx = header.index("Accuracy") + 1

    for df_idx in best_idx:
        excel_row = df_idx + 2
        cell = ws.cell(row=excel_row, column=accuracy_col_idx)
        cell.font = Font(bold=True)

    wb.save(output_path)


# =============================================================================
# MAIN
# =============================================================================
if __name__ == "__main__":
    df = build_dataframe_combined()
    print(df.head(80).to_string(index=False))
    print(f"\nFinal number of rows: {len(df)}")
    write_excel(df, OUTPUT_XLSX)
    print(f"Saved new file: {OUTPUT_XLSX}")